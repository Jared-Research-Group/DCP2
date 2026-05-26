import matplotlib.pyplot as plt
import openpyxl as xl
import numpy as np
import os

from tkinter import filedialog

def parse_keyence(f, plot=False):
    book = xl.load_workbook(f)
    sheet = book.active

    height = sheet.cell(row=16, column=2).value
    width  = sheet.cell(row=15, column=2).value
    step   = sheet.cell(row=13, column=2).value * (10**-3) # um to mm

    profile = np.zeros((height, width))

    for i in range(height):
        for j in range(width):
            profile[i][j] = sheet.cell(row=i + 25, column = j + 2).value

    x_pos = np.linspace(0, width * step, width)
    y_pos = np.linspace(0, height * step, height)

    x_pos, y_pos = np.meshgrid(x_pos, y_pos)

    if plot:
        fig, ax = plt.subplots(subplot_kw={'projection':'3d'})
        ax.plot_surface(x_pos, y_pos, profile, rstride=10, cstride=10, cmap='viridis')
        ax.set_aspect('equal')

        plt.show()

    return profile, x_pos, y_pos

def split_scan(dat):
    profile, x_pos, y_pos = dat

    uniques, counts = np.unique(profile, return_counts=True)
    profile -= uniques[np.argmax(counts)]

    grad = np.gradient(profile)

    fig, ax = plt.subplots(subplot_kw={'projection':'3d'})
    ax.plot_surface(x_pos, y_pos, grad[0], rstride=10, cstride=10, cmap='viridis')
    ax.set_xlabel('x position')
    ax.set_ylabel('y position')
    ax.set_zlabel('height gradient')
    plt.show()

    fig, ax = plt.subplots(subplot_kw={'projection':'3d'})
    ax.plot_surface(x_pos, y_pos, grad[1], rstride=10, cstride=10, cmap='viridis')
    ax.set_xlabel('x position')
    ax.set_ylabel('y position')
    ax.set_zlabel('height gradient')
    plt.show()



if __name__ == '__main__':
    keyence_excel = filedialog.askopenfilename()

    data = parse_keyence(keyence_excel, False)
    split_scan(data)